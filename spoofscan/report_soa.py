"""Statement of Applicability (SoA) workbook for an email audit.

Produces the spreadsheet an auditor usually asks for: one row per control, with
its status, the findings that justify it, the evidence and empty columns for
the owner and the target date, so it can be handed over as a work plan.

Requires openpyxl (``pip install openpyxl``); it is an optional dependency.
"""
from __future__ import annotations

from typing import Any, Dict, List

from . import catalog, scoring
from .models import Finding, Status

_HEADERS = {
    "en": ["Framework", "Control", "Title", "Family", "Status", "Findings",
           "Evidence / observations", "Owner", "Target date"],
    "es": ["Marco", "Control", "Titulo", "Familia", "Estado", "Hallazgos",
           "Evidencia / observaciones", "Responsable", "Fecha objetivo"],
}

_STATUS_FILL = {
    Status.COMPLIANT: "C6EFCE",
    Status.PARTIAL: "FFEB9C",
    Status.NON_COMPLIANT: "FFC7CE",
    Status.NOT_ASSESSED: "E7E6E6",
}


def write(
    path: str,
    posture: Dict[str, Any],
    findings: List[Finding],
    lang: str = "es",
    with_message: bool = False,
) -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - depends on environment
        raise RuntimeError(
            "openpyxl is required for the SoA workbook: pip install openpyxl"
        ) from exc

    statuses = scoring.control_status(findings, posture, with_message=with_message)
    headers = _HEADERS.get(lang, _HEADERS["es"])

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SoA" if lang == "en" else "Declaracion"

    title = (
        "Statement of Applicability - email authentication"
        if lang == "en" else
        "Declaracion de aplicabilidad - autenticacion de correo"
    )
    sheet["A1"] = title
    sheet["A1"].font = Font(size=14, bold=True)
    sheet["A2"] = (
        f"{posture.get('domain')} · {posture.get('collected_at')} · "
        f"score {scoring.score(findings, posture) if scoring.score(findings, posture) is not None else 'n/a'}"
    )
    sheet["A2"].font = Font(size=10, color="595959")

    header_row = 4
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    row = header_row + 1
    for ref in catalog.IN_SCOPE:
        entry = statuses[ref]
        control = entry["control"]
        state: Status = entry["status"]
        related: List[Finding] = entry["findings"]

        observations = " | ".join(
            f"{f.id}: {f.title.get(lang)}" for f in related
        ) or ("No automated finding" if lang == "en" else "Sin hallazgos automaticos")

        values = [
            control.framework,
            control.code,
            control.title.get(lang),
            control.family.get(lang),
            state.text(lang),
            ", ".join(f.id for f in related),
            observations,
            "",
            "",
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=column in (3, 4, 7))
        sheet.cell(row=row, column=5).fill = PatternFill(
            "solid", fgColor=_STATUS_FILL[state]
        )
        row += 1

    widths = [11, 12, 34, 30, 15, 16, 56, 18, 14]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{row - 1}"

    # Second sheet: the raw findings, for traceability of the SoA statuses.
    detail = workbook.create_sheet("Findings" if lang == "en" else "Hallazgos")
    detail_headers = (
        ["ID", "Severity", "Title", "Detail", "Recommendation", "Controls",
         "Reference", "Evidence"]
        if lang == "en" else
        ["ID", "Severidad", "Titulo", "Detalle", "Recomendacion", "Controles",
         "Referencia", "Evidencia"]
    )
    for column, header in enumerate(detail_headers, start=1):
        cell = detail.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")

    from .models import sort_findings

    for index, finding in enumerate(sort_findings(findings), start=2):
        cells = [
            finding.id,
            finding.severity.label,
            finding.title.get(lang),
            finding.detail.get(lang),
            finding.recommendation.get(lang),
            catalog.pretty(finding.controls),
            finding.reference or "",
            finding.evidence or "",
        ]
        for column, value in enumerate(cells, start=1):
            cell = detail.cell(row=index, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=column in (3, 4, 5))
    for index, width in enumerate([10, 12, 38, 70, 62, 30, 16, 30], start=1):
        detail.column_dimensions[get_column_letter(index)].width = width
    detail.freeze_panes = "A2"

    workbook.save(path)
    return path
